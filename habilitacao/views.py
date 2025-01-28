from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth import authenticate,login,logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from gutilizador.models import *
from .models import *
from .forms import *
from django.db import transaction
from django.utils import timezone
now = timezone.now()
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.db.models import Q
from django.db.models.functions import ExtractYear
# Create your views here.
@login_required
def index(request):
	objects = DashboardHabilitacao.objects.all()
	objects_nivel_hab = Nivel_habilitacao.objects.filter(controlo_estado__isnull=True)
	groups = request.user.groups.all()
	group_names = [group.name for group in groups]
	objects1 = ViewMonitorHabilitacao.objects.all()
	hab_years = ViewMonitorHabilitacao.objects.annotate(year=ExtractYear('data_fim_habilitacao')).values_list('year', flat=True).distinct()
	# print("hab_years:",hab_years)
	context = {
		"title":"Dashboard Habilitação Académica",
		"legend":"Dashboard Habilitação Académica",
		"active_dash_ad":'active',
		"page":'dash_ad',
		"objects":objects,
		"hab_years":hab_years,
		"objects_nivel_hab":objects_nivel_hab,
	}
	return render(request,'habilitacao_ac/dash.html',context)

@login_required
def monitor_habilitacao_data(request):
	start = int(request.GET.get('start', 0))
	length = int(request.GET.get('length', 10))
	search_value = request.GET.get('search[value]', '')
	draw = int(request.GET.get('draw', 1))
	nome_fun = request.GET.get('nome_fun', '')
	id_sigap = request.GET.get('id_sigap', '')
	sexo = request.GET.get('sexo', '')
	nivel_habilitacao = request.GET.get('nivel_habilitacao', '')
	curso = request.GET.get('curso', '')
	entidade = request.GET.get('entidade', '')
	ano_hab = request.GET.get('ano_hab', '')
	print("ano_hab:",ano_hab)

	# Handle ordering
	order_column_index = int(request.GET.get('order[0][column]', 0))
	order_direction = request.GET.get('order[0][dir]', 'asc')

	column_mapping = {
		0: 'id',  # Assuming first column is a combination, you can use 'id' or another field
		1: 'nome',  # Corresponds to 'nome_do_funcionario'
		2: 'sexo',
		3: 'id_sigap',
		4: 'data_inicio_habilitacao',
		5: 'data_fim_habilitacao',
		6: 'nivel_habilitacao',
		7: 'curso',
		8: 'entidade',
	}
	order_column = column_mapping.get(order_column_index, 'id')
	if order_direction == 'desc':
		order_column = f'-{order_column}'

	# Query the database
	data_objects = ViewMonitorHabilitacao.objects.all()
	allTotal = data_objects.count()
	if nome_fun:
		data_objects = data_objects.filter(nome__icontains=nome_fun)
	if id_sigap:
		data_objects = data_objects.filter(id_sigap__icontains=id_sigap)
	if sexo:
		data_objects = data_objects.filter(sexo=sexo)
	if nivel_habilitacao:
		data_objects = data_objects.filter(nivel_habilitacao__icontains=nivel_habilitacao)
	if curso:
		data_objects = data_objects.filter(curso__icontains=curso)
	if entidade:
		data_objects = data_objects.filter(entidade__icontains=entidade)
	if ano_hab:
		data_objects = data_objects.filter(data_fim_habilitacao__year=ano_hab)

	# Apply ordering
	data_objects = data_objects.order_by(order_column)

	# Pagination
	paginator = Paginator(data_objects, length)
	page = paginator.get_page(start // length + 1)
	totalFiltered = data_objects.count()


	# Prepare data for DataTables
	data = []
	for idx, dados in enumerate(page.object_list, start=1):
		data.append({
        	'id': dados.id,
        	'nome_do_funcionario': dados.nome,
        	'sexo': dados.sexo,
        	'id_sigap': dados.id_sigap,
        	'data_inicio_habilitacao': dados.data_inicio_habilitacao.strftime('%d/%m/%Y') if dados.data_inicio_habilitacao else "",
        	'data_fim_habilitacao': dados.data_fim_habilitacao.strftime('%d/%m/%Y') if dados.data_fim_habilitacao else "",
        	'nivel_habilitacao': dados.nivel_habilitacao,
        	'curso': dados.curso,
        	'entidade': dados.entidade,
    	})

	response = {
		"draw": draw,
		"recordsTotal": paginator.count,
		"recordsFiltered": paginator.count,
		"totalFiltered": totalFiltered,
		"allTotal": allTotal,
		"data": data,
	}
	return JsonResponse(response)

from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

@login_required
def export_monitor_habilitacao_to_excel(request):
    # Get filters from the request
    nome_fun = request.GET.get('nome_fun', '')
    id_sigap = request.GET.get('id_sigap', '')
    sexo = request.GET.get('sexo', '')
    nivel_habilitacao = request.GET.get('nivel_habilitacao', '')
    curso = request.GET.get('curso', '')
    entidade = request.GET.get('entidade', '')
    ano_hab = request.GET.get('ano_hab', '')

    # Filter the data based on the filters
    data_objects = ViewMonitorHabilitacao.objects.all()
    if nome_fun:
        data_objects = data_objects.filter(nome__icontains=nome_fun)
    if id_sigap:
        data_objects = data_objects.filter(id_sigap__icontains=id_sigap)
    if sexo:
        data_objects = data_objects.filter(sexo=sexo)
    if nivel_habilitacao:
        data_objects = data_objects.filter(nivel_habilitacao__icontains=nivel_habilitacao)
    if curso:
        data_objects = data_objects.filter(curso__icontains=curso)
    if entidade:
        data_objects = data_objects.filter(entidade__icontains=entidade)
    if ano_hab:
        data_objects = data_objects.filter(data_fim_habilitacao__year=ano_hab)

    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Monitoramento Habilitação"

    # Add headers
    headers = [
        "Nome do Funcionário", "Sexo", "ID Sigap",
        "Data Início Habilitação", "Data Fim Habilitação",
        "Nível Habilitação", "Curso", "Entidade"
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Add data rows
    for row_num, obj in enumerate(data_objects, start=2):
        sheet[f"A{row_num}"] = obj.nome
        sheet[f"B{row_num}"] = obj.sexo
        sheet[f"C{row_num}"] = obj.id_sigap
        sheet[f"D{row_num}"] = obj.data_inicio_habilitacao.strftime('%d/%m/%Y') if obj.data_inicio_habilitacao else ""
        sheet[f"E{row_num}"] = obj.data_fim_habilitacao.strftime('%d/%m/%Y') if obj.data_fim_habilitacao else ""
        sheet[f"F{row_num}"] = obj.nivel_habilitacao
        sheet[f"G{row_num}"] = obj.curso
        sheet[f"H{row_num}"] = obj.entidade

    # Set the response for file download
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="monitoramento_habilitacao.xlsx"'
    workbook.save(response)
    return response


@login_required
def DetalhoFuncionarioListaHab(request,id):
	objects = Funcionarios.objects.get(id=id)
	objects1 = Habilitacao_academica.objects.filter(controlo_estado__isnull=True).filter(funcionario=objects).order_by('-data_habilitacao')
	context = {
	'link_antes': [{'link_name':"lista-funcionarios",'link_text':"Lista Funcionarios"},{'link_name':"dashboard-habilitacao",'link_text':"Dashboard Habilitação Académica"}],
		"title":f"Dados Habilitação {objects.nome}",
		"legend":f"Dados Habilitação {objects.nome}",
		"objects":objects,
		"objects1":objects1,
		"page":'detalho_funcionario_lista_hab',
		"active_det_fun_list_hab":'active',
	}
	return render(request, 'habilitacao_ac/detalho_hab_funcionario.html', context)


@login_required
def DetalhoFuncionarioRegistaHab(request,id):
	objects = Funcionarios.objects.get(id=id)
	if request.method == "POST":
		form = Habilitacao_academicaForm(request.POST, request.FILES)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.funcionario = objects
				instance.inserido_por = request.user
				instance.save()
				messages.success(request, f'Nova Ficha de Habilitação foi inserido com sucesso!')
				return redirect('DetalhoFuncionarioListaHab',objects.id)
	else:
		form = Habilitacao_academicaForm()
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"dashboard-habilitacao",'link_text':"Dashboard Habilitação Académica"},
		{'link_name':"DetalhoFuncionarioListaHab",'link_param':f"{objects.id}",'link_text':f"Dados Habilitação {objects.nome}"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Inserir Ficha de Habilitação",
		"legend":"Inserir Ficha de Habilitação",
		"page":"form_hab_fun",
		"active_inserir_hab_fun":"active",
		"objects":objects,
		"form":form,
	}
	return render(request, 'funcionario/form_tab.html', context)

@login_required
def RegistaNovoHabFuncionario(request):
	# objects = Funcionarios.objects.get(id=id)
	if request.method == "POST":
		form = Habilitacao_academicaFunForm(request.POST, request.FILES)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				fun = Funcionarios.objects.get(id=instance.funcionario.id)
				instance.inserido_por = request.user
				# instance.save()
				messages.success(request, f'Nova Ficha de Habilitação foi inserido com sucesso!')
				return redirect('DetalhoFuncionarioListaHab',fun.id)
	else:
		form = Habilitacao_academicaFunForm()
	link_antes = [
		{'link_name':"dashboard-habilitacao",'link_param':"",'link_text':f"Dashboard Habilitação Académica"},
		# {'link_name':"DetalhoFuncionarioListaHab",'link_param':f"{fun.id}",'link_text':f"Dados Habilitação {objects.nome}"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Inserir Ficha de Habilitação",
		"legend":"Inserir Ficha de Habilitação",
		"page":"form_hab_fun",
		"active_inserir_hab_fun":"active",
		# "objects":objects,
		"form":form,
	}
	return render(request, 'habilitacao_ac/forms.html', context)

@login_required
def DetalhoFuncionarioDesabilitarHab(request,id):
	objects1 = Habilitacao_academica.objects.get(id=id)
	objects = Funcionarios.objects.get(id=objects1.funcionario.id)
	with transaction.atomic():
		objects1.controlo_estado = 'Desabilitar'
		objects1.alterado_por = request.user.username
		objects1.alterado_em = now
		objects1.save()
		messages.success(request, f'Ficha de Habilitação foi desabilitar com sucesso!')
		return redirect('DetalhoFuncionarioListaHab',objects.id)

@login_required
def DetalhoFuncionarioAlteraHab(request,id):
	objects1 = Habilitacao_academica.objects.get(id=id)
	objects = Funcionarios.objects.get(id=objects1.funcionario.id)
	if request.method == "POST":
		form = Habilitacao_academicaForm(request.POST, request.FILES,instance=objects1)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.funcionario = objects
				instance.alterado_por = request.user.username
				instance.alterado_em = now
				instance.save()
				messages.success(request, f'Ficha de Habilitação foi alterado com sucesso!')
				return redirect('DetalhoFuncionarioListaHab',objects.id)
	else:
		form = Habilitacao_academicaForm(instance=objects1)
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"dashboard-habilitacao",'link_text':"Dashboard Habilitação Académica"},
		{'link_name':"DetalhoFuncionarioListaHab",'link_param':f"{objects.id}",'link_text':f"Dados Habilitação {objects.nome}"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Alterar Ficha de Habilitação",
		"legend":"Alterar Ficha de Habilitação",
		"page":"form_hab_fun",
		"active_alterar_hab_fun":"active",
		"objects":objects,
		"form":form,
	}
	return render(request, 'funcionario/form_tab.html', context)



@login_required
def nivelHabilitacao(request):
	objects = Nivel_habilitacao.objects.all().order_by('ordem')
	groups = request.user.groups.all()
	group_names = [group.name for group in groups]
	context = {
		"title":"Dados Nivel Habilitação",
		"legend":"Dados Nivel Habilitação",
		"active_nivel_hab":'active',
		"page":'nivel_hab',
		"objects":objects,
	}
	return render(request,'parametro/nivel_hab.html',context)

@login_required
def nivelHabilitacaoAdd(request):
	if request.method == "POST":
		form = Nivel_habilitacaoForm(request.POST, request.FILES)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.inserido_por = request.user
				instance.save()
				messages.success(request, f'Novo Dados de Nivel Habilitação foi inserido com sucesso!')
				return redirect('nivelHabilitacao')
	else:
		form = Nivel_habilitacaoForm()
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"nivelHabilitacao",'link_param':"",'link_text':f"Dados Nivel Habilitação"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Inserir Nivel de Habilitação",
		"legend":"Inserir Nivel de Habilitação",
		"page":"form_nivel_hab",
		"active_inserir_nivel_hab":"active",
		"form":form,
	}
	return render(request, 'habilitacao_ac/forms.html', context)

@login_required
def nivelHabilitacaoAltera(request,id):
	objects = Nivel_habilitacao.objects.get(id=id)
	if request.method == "POST":
		form = Nivel_habilitacaoForm(request.POST, request.FILES,instance=objects)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.alterado_por = request.user.username
				instance.alterado_em = now
				instance.save()
				messages.success(request, f'Dados de Nivel Habilitação foi alterado com sucesso!')
				return redirect('nivelHabilitacao')
	else:
		form = Nivel_habilitacaoForm(instance=objects)
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"nivelHabilitacao",'link_param':"",'link_text':f"Dados Nivel Habilitação"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Alterar Nivel de Habilitação",
		"legend":"Alterar Nivel de Habilitação",
		"page":"form_nivel_hab",
		"active_alterar_nivel_hab":"active",
		"form":form,
	}
	return render(request, 'habilitacao_ac/forms.html', context)

@login_required
def EntidadeHabilitacao(request):
	objects = Entidade_habilitacao.objects.all()
	groups = request.user.groups.all()
	group_names = [group.name for group in groups]
	context = {
		"title":"Dados Entidade Habilitação",
		"legend":"Dados Entidade Habilitação",
		"active_entidade_hab":'active',
		"page":'entidade_hab',
		"objects":objects,
	}
	return render(request,'parametro/entidade_hab.html',context)

@login_required
def EntidadeHabilitacaoAdd(request):
	if request.method == "POST":
		form = Entidade_habilitacaoForm(request.POST, request.FILES)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.inserido_por = request.user
				instance.save()
				messages.success(request, f'Novo Dados de Entidade Habilitação foi inserido com sucesso!')
				return redirect('EntidadeHabilitacao')
	else:
		form = Entidade_habilitacaoForm()
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"EntidadeHabilitacao",'link_param':"",'link_text':f"Dados Entidade Habilitação"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Inserir Entidade de Habilitação",
		"legend":"Inserir Entidade de Habilitação",
		"page":"form_entidade_hab",
		"active_inserir_entidade_hab":"active",
		"form":form,
	}
	return render(request, 'habilitacao_ac/forms.html', context)

@login_required
def EntidadeHabilitacaoAltera(request,id):
	objects = Entidade_habilitacao.objects.get(id=id)
	if request.method == "POST":
		form = Entidade_habilitacaoForm(request.POST, request.FILES,instance=objects)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.alterado_por = request.user.username
				instance.alterado_em = now
				instance.save()
				messages.success(request, f'Dados de Entidade Habilitação foi alterado com sucesso!')
				return redirect('EntidadeHabilitacao')
	else:
		form = Entidade_habilitacaoForm(instance=objects)
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"EntidadeHabilitacao",'link_param':"",'link_text':f"Dados Entidade Habilitação"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Alterar Entidade de Habilitação",
		"legend":"Alterar Entidade de Habilitação",
		"page":"form_entidade_hab",
		"active_alterar_entidade_hab":"active",
		"form":form,
	}
	return render(request, 'habilitacao_ac/forms.html', context)

@login_required
def CursoHabilitacao(request):
	objects = Curso_habilitacao.objects.all()
	groups = request.user.groups.all()
	group_names = [group.name for group in groups]
	context = {
		"title":"Dados Curso de Habilitação",
		"legend":"Dados Curso de Habilitação",
		"active_curso_hab":'active',
		"page":'curso_hab',
		"objects":objects,
	}
	return render(request,'parametro/curso_hab.html',context)

@login_required
def CursoHabilitacaoAdd(request):
	if request.method == "POST":
		form = Curso_habilitacaoForm(request.POST, request.FILES)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.inserido_por = request.user
				instance.save()
				messages.success(request, f'Novo Dados de Curso de Habilitação foi inserido com sucesso!')
				return redirect('CursoHabilitacao')
	else:
		form = Curso_habilitacaoForm()
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"CursoHabilitacao",'link_param':"",'link_text':f"Dados Curso Habilitação"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Inserir Curso de Habilitação",
		"legend":"Inserir Curso",
		"page":"form_curso_hab",
		"active_inserir_curso_hab":"active",
		"form":form,
	}
	return render(request, 'habilitacao_ac/forms.html', context)

@login_required
def CursoHabilitacaoAltera(request,id):
	objects = Curso_habilitacao.objects.get(id=id)
	if request.method == "POST":
		form = Curso_habilitacaoForm(request.POST, request.FILES,instance=objects)
		if form.is_valid():
			with transaction.atomic():
				instance = form.save(commit=False)
				instance.alterado_por = request.user.username
				instance.alterado_em = now
				instance.save()
				messages.success(request, f'Dados de Curso de Habilitação foi alterado com sucesso!')
				return redirect('CursoHabilitacao')
	else:
		form = Curso_habilitacaoForm(instance=objects)
	link_antes = [
		{'link_name':"lista-funcionarios",'link_param':"",'link_text':f"Lista Funcionarios"},
		{'link_name':"CursoHabilitacao",'link_param':"",'link_text':f"Dados Curso Habilitação"}
	]
	context = {
		'link_antes':link_antes,
		"title":"Alterar Curso Habilitação",
		"legend":"Alterar Curso",
		"page":"form_curso_hab",
		"active_alterar_curso_hab":"active",
		"form":form,
	}
	return render(request, 'habilitacao_ac/forms.html', context)